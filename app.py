from flask import Flask, render_template, request, jsonify,send_from_directory
import json
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from flask import Flask, request, abort
from linebot.v3 import (WebhookHandler)
from linebot.v3.exceptions import (InvalidSignatureError)
from linebot.v3.messaging import (Configuration, ApiClient,MessagingApi,ReplyMessageRequest,TextMessage)
from linebot.v3.webhooks import (MessageEvent,TextMessageContent)
from linebot.v3.messaging.models import (PushMessageRequest,TemplateMessage,ButtonsTemplate,PostbackAction,MessageAction,URIAction)
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from dotenv import load_dotenv
import math
import os
from collections import defaultdict
import pyodbc
ENV = './.env' 
load_dotenv(dotenv_path=ENV)

app = Flask(__name__)

# ===== LINE 設定 =====
CHANNEL_ACCESS_TOKEN = os.getenv('CHANNEL_ACCESS_TOKEN')  # Messaging API Channel Access Token
CHANNEL_SECRET = os.getenv('CHANNEL_SECRET')
# ===== 路徑 設定 =====
TEMP='temp'
PNG='static/img'
FOLDER='static/file'
app.config['TEMP'] = TEMP
app.config['PNG'] = PNG
app.config['FOLDER'] = FOLDER
configuration = Configuration(access_token=CHANNEL_ACCESS_TOKEN)
handler = WebhookHandler(CHANNEL_SECRET)
api_client = ApiClient(configuration)
line_bot_api = MessagingApi(api_client)
# ===== 全域資料 =====
last_setting={"hour": 9, "minute": 0} 
settings = {"hour": 9, "minute": 0}  # 每日推送時間
def safe_float(val):
    return float(val) if val is not None else 0.0
def safe_int(val):
    return int(val) if val is not None else 0
def getdailydata(User,Date):
    with open("store.json", "r", encoding="utf-8") as f:
        stores = json.load(f)
        

    with open("permissions.json", "r", encoding="utf-8") as f:
        permissions = json.load(f)
    depart=[]
    for per in permissions:
        if User==per['user_id']:
            if per['departments'][0]=='all':
                depart = [stor['value'] for stor in stores]
            else:
                depart=per['departments']
              
                break
    load_dotenv()
    Daily_HOST = os.getenv('Daily_HOST')
    Daily_password = os.getenv('Daily_password')
    Daily_uid=os.getenv('Daily_uid')
    Daily_name=os.getenv('Daily_name')
    conn = pyodbc.connect(
        f"DRIVER={{ODBC Driver 17 for SQL Server}};"
        f"SERVER={Daily_HOST};"
        f"DATABASE={Daily_name};"
        f"UID={Daily_uid};"
        f"PWD={Daily_password};"
        "Trusted_Connection=no;"
    )
    cursor = conn.cursor()
    Date = datetime.strptime(Date, "%Y-%m-%d")
    DateMonth = Date.strftime("%Y%m")
    first_month = datetime(Date.year, 1, 1)
    FirstMonth=first_month.strftime("%Y%m")
    try:
        Date_last_year = Date.replace(year=Date.year - 1)
        DateMonth_last_year = Date_last_year.strftime("%Y%m")
        first_month_last_year = datetime(Date_last_year.year, 1, 1)
        FirstMonth_last_year=first_month_last_year.strftime("%Y%m")
    except ValueError:
        # 如果是 2/29 會錯誤，可以退一天
        Date_last_year = Date.replace(year=Date.year - 1, day=28)
        DateMonth_last_year = Date_last_year.strftime("%Y%m")
        first_month_last_year = datetime(Date_last_year.year, 1, 1)
        FirstMonth_last_year=first_month_last_year.strftime("%Y%m")
    cursor.execute("SELECT store_id, total_amt ,total_customer,sales_count,DATE FROM kingza_api.dbo.SalesAggregate WHERE DATE = ?", (Date,))
    dsc = cursor.fetchall()
    cursor.execute("SELECT store_id, total_amt ,total_customer,sales_count,DATE FROM kingza_api.dbo.SalesAggregate WHERE DATE = ?", (Date_last_year,))
    dsp=cursor.fetchall()
    cursor.execute("SELECT store_id, total_amt ,total_customer,sales_count,Month FROM kingza_api.dbo.SalesAggregateByMonth WHERE Month = ?", (DateMonth,))
    msc=cursor.fetchall()
    cursor.execute("SELECT store_id, total_amt ,total_customer,sales_count,Month FROM kingza_api.dbo.SalesAggregateByMonth WHERE Month = ?", (DateMonth_last_year,))
    msp=cursor.fetchall()
    cursor.execute("""
        SELECT store_id,
            SUM(total_amt) AS total_amt,
            SUM(total_customer) AS total_customer,
            SUM(sales_count) AS sales_count
        FROM kingza_api.dbo.SalesAggregateByMonth
        WHERE Month >= ? AND Month <= ?
        GROUP BY store_id;
    """, (FirstMonth, DateMonth))
    ysc=cursor.fetchall()
    cursor.execute("""
        SELECT store_id,
            SUM(total_amt) AS total_amt,
            SUM(total_customer) AS total_customer,
            SUM(sales_count) AS sales_count
        FROM kingza_api.dbo.SalesAggregateByMonth
        WHERE Month >= ? AND Month <= ?
        GROUP BY store_id;
    """, (FirstMonth_last_year, DateMonth_last_year))
    ysp=cursor.fetchall()
    store_map = {s['value']: s for s in stores}
    dsc_map = {row[0]: row for row in dsc}
    dsp_map = {row[0]: row for row in dsp}
    msc_map = {row[0]: row for row in msc}
    msp_map = {row[0]: row for row in msp}
    ysc_map = {row[0]: row for row in ysc}
    ysp_map = {row[0]: row for row in ysp}
    result=[]
    for store_id in depart :
        store_info = store_map.get(store_id, {})
        dsc_row = dsc_map.get(store_id, (store_id, 0, 0, 0, None))
        dsp_row = dsp_map.get(store_id, (store_id, 0, 0, 0, None))
        msc_row = msc_map.get(store_id, (store_id, 0, 0, 0, None))
        msp_row = msp_map.get(store_id, (store_id, 0, 0, 0, None))
        ysc_row = ysc_map.get(store_id, (store_id, 0, 0, 0, None))
        ysp_row = ysp_map.get(store_id, (store_id, 0, 0, 0, None))
        result.append({
            "store_id": store_id,
            "store_name": store_info.get("name", ""),
            "dept": store_info.get("dept", ""),
            
            "dsc_total_amt": safe_float(dsc_row[1]),
            "dsc_total_customer": safe_int(dsc_row[2]),
            "dsc_sales_count": safe_int(dsc_row[3]),
            
            
            "dsp_total_amt": safe_float(dsp_row[1]),
            "dsp_total_customer": safe_int(dsp_row[2]),
            "dsp_sales_count": safe_int(dsp_row[3]),
            

            "msc_total_amt": safe_float(msc_row[1]),
            "msc_total_customer": safe_int(msc_row[2]),
            "msc_sales_count": safe_int(msc_row[3]),
            

            "msp_total_amt": safe_float(msp_row[1]),
            "msp_total_customer": safe_int(msp_row[2]),
            "msp_sales_count": safe_int(msp_row[3]),
            

            "ysc_total_amt": safe_float(ysc_row[1]),
            "ysc_total_customer": safe_int(ysc_row[2]),
            "ysc_sales_count": safe_int(ysc_row[3]),
            

            "ysp_total_amt": safe_float(ysp_row[1]),
            "ysp_total_customer": safe_int(ysp_row[2]),
            "ysp_sales_count": safe_int(ysp_row[3]),
            
        })
    data = []
    D_total={
            
            "dsc_total_amt": 0.0,
            "dsc_total_customer": 0,
            "dsc_sales_count": 0,
            
            
            "dsp_total_amt": 0,
            "dsp_total_customer": 0,
            "dsp_sales_count": 0,
            

            "msc_total_amt": 0,
            "msc_total_customer": 0,
            "msc_sales_count": 0,
            

            "msp_total_amt": 0,
            "msp_total_customer": 0,
            "msp_sales_count": 0,
            

            "ysc_total_amt": 0,
            "ysc_total_customer": 0,
            "ysc_sales_count": 0,
            

            "ysp_total_amt": 0,
            "ysp_total_customer": 0,
            "ysp_sales_count": 0,
            
        }
    totals = {}
    for r in result:
    # 先取得店名或其他欄位
        row = [
            r["store_name"],                 # 店名
            r["dept"],                       # 區經理
            
            r["dsc_total_amt"],              # dsc 總額
            r["dsp_total_amt"],              # dsp 總額
            r["dsc_total_amt"] / r["dsp_total_amt"] if r["dsp_total_amt"] else None,  # 比例
            
            r["dsc_total_customer"],         # dsc 總客數
            r["dsp_total_customer"],         # dsp 總客數
            r["dsc_total_customer"] / r["dsp_total_customer"] if r["dsp_total_customer"] else None,  # 比例
            
            r["dsc_sales_count"],            # dsc 銷售筆數
            r["dsp_sales_count"],            # dsp 銷售筆數
            r["dsc_sales_count"] / r["dsp_sales_count"] if r["dsp_sales_count"] else None,  # 比例

            #################################################################
            r["msc_total_amt"],              # msc 總額
            r["msp_total_amt"],              # msp 總額
            r["msc_total_amt"] / r["msp_total_amt"] if r["msp_total_amt"] else None,  # 比例
            
            r["msc_total_customer"],         # msc 總客數
            r["msp_total_customer"],         # msp 總客數
            r["msc_total_customer"] / r["msp_total_customer"] if r["msp_total_customer"] else None,  # 比例
            
            r["msc_sales_count"],            # msc 銷售筆數
            r["msp_sales_count"],            # msp 銷售筆數
            r["msc_sales_count"] / r["msp_sales_count"] if r["msp_sales_count"] else None,  # 比例
            #################################################################################
            r["ysc_total_amt"],              # dsc 總額
            r["ysp_total_amt"],              # dsp 總額
            r["ysc_total_amt"] / r["ysp_total_amt"] if r["ysp_total_amt"] else None,  # 比例
            
            r["ysc_total_customer"],         # dsc 總客數
            r["ysp_total_customer"],         # dsp 總客數
            r["ysc_total_customer"] / r["ysp_total_customer"] if r["ysp_total_customer"] else None,  # 比例
            
            r["ysc_sales_count"],            # dsc 銷售筆數
            r["ysp_sales_count"],            # dsp 銷售筆數
            r["ysc_sales_count"] / r["ysp_sales_count"] if r["ysp_sales_count"] else None,  # 比例
        ]
        key = f"{r["store_name"][0]}Total" # 取第一個字
        
        if key not in totals:
            totals[key] = {
            
                "dsc_total_amt": 0.0,
                "dsc_total_customer": 0,
                "dsc_sales_count": 0,
                
                
                "dsp_total_amt": 0,
                "dsp_total_customer": 0,
                "dsp_sales_count": 0,
                

                "msc_total_amt": 0,
                "msc_total_customer": 0,
                "msc_sales_count": 0,
                

                "msp_total_amt": 0,
                "msp_total_customer": 0,
                "msp_sales_count": 0,
                

                "ysc_total_amt": 0,
                "ysc_total_customer": 0,
                "ysc_sales_count": 0,
                

                "ysp_total_amt": 0,
                "ysp_total_customer": 0,
                "ysp_sales_count": 0,
                
            }
        totals[key]["dsc_total_amt"] += r["dsc_total_amt"]
        totals[key]["dsp_total_amt"] += r["dsp_total_amt"]
        totals[key]["dsc_total_customer"] += r["dsc_total_customer"]
        totals[key]["dsp_total_customer"] += r["dsp_total_customer"]
        totals[key]["dsc_sales_count"] += r["dsc_sales_count"]
        totals[key]["dsp_sales_count"] += r["dsp_sales_count"]
        ##
        totals[key]["msc_total_amt"] += r["msc_total_amt"]
        totals[key]["msp_total_amt"] += r["msp_total_amt"]
        totals[key]["msc_total_customer"] += r["msc_total_customer"]
        totals[key]["msp_total_customer"] += r["msp_total_customer"]
        totals[key]["msc_sales_count"] += r["msc_sales_count"]
        totals[key]["msp_sales_count"] += r["msp_sales_count"]
        ##
        totals[key]["ysc_total_amt"] += r["ysc_total_amt"]
        totals[key]["ysp_total_amt"] += r["ysp_total_amt"]
        totals[key]["ysc_total_customer"] += r["ysc_total_customer"]
        totals[key]["ysp_total_customer"] += r["ysp_total_customer"]
        totals[key]["ysc_sales_count"] += r["ysc_sales_count"]
        totals[key]["ysp_sales_count"] += r["ysp_sales_count"]
        D_total["dsc_total_amt"] = r["dsc_total_amt"] + D_total["dsc_total_amt"]
        D_total["dsp_total_amt"] = r["dsp_total_amt"] + D_total["dsp_total_amt"]
        D_total["dsc_total_customer"] = r["dsc_total_customer"] + D_total["dsc_total_customer"]
        D_total["dsp_total_customer"] = r["dsp_total_customer"] + D_total["dsp_total_customer"]
        D_total["dsc_sales_count"] = r["dsc_sales_count"] + D_total["dsc_sales_count"]
        D_total["dsp_sales_count"] = r["dsp_sales_count"] + D_total["dsp_sales_count"]
        ##
        D_total["msc_total_amt"] = r["msc_total_amt"] + D_total["msc_total_amt"]
        D_total["msp_total_amt"] = r["msp_total_amt"] + D_total["msp_total_amt"]
        D_total["msc_total_customer"] = r["msc_total_customer"] + D_total["msc_total_customer"]
        D_total["msp_total_customer"] = r["msp_total_customer"] + D_total["msp_total_customer"]
        D_total["msc_sales_count"] = r["msc_sales_count"] + D_total["msc_sales_count"]
        D_total["msp_sales_count"] = r["msp_sales_count"] + D_total["msp_sales_count"]
        ##
        D_total["ysc_total_amt"] = r["ysc_total_amt"] + D_total["ysc_total_amt"]
        D_total["ysp_total_amt"] = r["ysp_total_amt"] + D_total["ysp_total_amt"]
        D_total["ysc_total_customer"] = r["ysc_total_customer"] + D_total["ysc_total_customer"]
        D_total["ysp_total_customer"] = r["ysp_total_customer"] + D_total["ysp_total_customer"]
        D_total["ysc_sales_count"] = r["ysc_sales_count"] + D_total["ysc_sales_count"]
        D_total["ysp_sales_count"] = r["ysp_sales_count"] + D_total["ysp_sales_count"]
        ##
        data.append(row)
        data.sort(key=lambda x: x[1])
    # print(key)
    brand_data=[]
    for total in totals:
        
        brd=[
            total,                 # 店名
            '', 
            totals[total]["dsc_total_amt"],
            totals[total]["dsp_total_amt"],
            totals[total]["dsc_total_amt"] / totals[total]["dsp_total_amt"] if totals[total]["dsp_total_amt"] else None,
            totals[total]["dsc_total_customer"],
            totals[total]["dsp_total_customer"],
            totals[total]["dsc_total_customer"] / totals[total]["dsp_total_customer"] if totals[total]["dsp_total_customer"] else None,
            totals[total]["dsc_sales_count"],
            totals[total]["dsp_sales_count"],
            totals[total]["dsc_sales_count"] / totals[total]["dsp_sales_count"] if totals[total]["dsp_sales_count"] else None,
            totals[total]["msc_total_amt"],
            totals[total]["msp_total_amt"],
            totals[total]["msc_total_amt"] / totals[total]["msp_total_amt"] if totals[total]["msp_total_amt"] else None,
            totals[total]["msc_total_customer"],
            totals[total]["msp_total_customer"],
            totals[total]["msc_total_customer"] / totals[total]["msp_total_customer"] if totals[total]["msp_total_customer"] else None,
            totals[total]["msc_sales_count"],
            totals[total]["msp_sales_count"],
            totals[total]["msc_sales_count"] / totals[total]["msp_sales_count"] if totals[total]["msp_sales_count"] else None,
            totals[total]["ysc_total_amt"],
            totals[total]["ysp_total_amt"],
            totals[total]["ysc_total_amt"] / totals[total]["ysp_total_amt"] if totals[total]["ysp_total_amt"] else None,
            totals[total]["ysc_total_customer"],
            totals[total]["ysp_total_customer"],
            totals[total]["ysc_total_customer"] / totals[total]["ysp_total_customer"] if totals[total]["ysp_total_customer"] else None,
            totals[total]["ysc_sales_count"],
            totals[total]["ysp_sales_count"],
            totals[total]["ysc_sales_count"] / totals[total]["ysp_sales_count"] if totals[total]["ysp_sales_count"] else None
        ]
        brand_data.append(brd)
    for i in range(len(brand_data)): 
        data.insert(0, brand_data[i])
    D_TOTAL_DATA=[
        'Total',                 # 店名
        '',     
        D_total["dsc_total_amt"],
        D_total["dsp_total_amt"],
        D_total["dsc_total_amt"] / D_total["dsp_total_amt"] if D_total["dsp_total_amt"] else None,
        D_total["dsc_total_customer"],
        D_total["dsp_total_customer"],
        D_total["dsc_total_customer"] / D_total["dsp_total_customer"] if D_total["dsp_total_customer"] else None,
        D_total["dsc_sales_count"],
        D_total["dsp_sales_count"],
        D_total["dsc_sales_count"] / D_total["dsp_sales_count"] if D_total["dsp_sales_count"] else None,
        D_total["msc_total_amt"],
        D_total["msp_total_amt"],
        D_total["msc_total_amt"] / D_total["msp_total_amt"] if D_total["msp_total_amt"] else None,
        D_total["msc_total_customer"],
        D_total["msp_total_customer"],
        D_total["msc_total_customer"] / D_total["msp_total_customer"] if D_total["msp_total_customer"] else None,
        D_total["msc_sales_count"],
        D_total["msp_sales_count"],
        D_total["msc_sales_count"] / D_total["msp_sales_count"] if D_total["msp_sales_count"] else None,
        D_total["ysc_total_amt"],
        D_total["ysp_total_amt"],
        D_total["ysc_total_amt"] / D_total["ysp_total_amt"] if D_total["ysp_total_amt"] else None,
        D_total["ysc_total_customer"],
        D_total["ysp_total_customer"],
        D_total["ysc_total_customer"] / D_total["ysp_total_customer"] if D_total["ysp_total_customer"] else None,
        D_total["ysc_sales_count"],
        D_total["ysp_sales_count"],
        D_total["ysc_sales_count"] / D_total["ysp_sales_count"] if D_total["ysp_sales_count"] else None,
    ]
    
    # data[0].append(D_TOTAL_DATA)
    # print(data)
    data.insert(0, D_TOTAL_DATA)
    return data      
def update_job():
    """檢查設定是否改變，更新排程"""
    global current_job, last_setting
    with open("settings.json", "r", encoding="utf-8") as f:
        settings = json.load(f)
    setting=settings[0]
    hour = setting.get("hour", 9)
    minute = setting.get("minute", 0)
    # 刪掉舊 job
    if current_job:
        scheduler.remove_job(current_job.id)
    # 建立新 job
    trigger = CronTrigger(hour=hour, minute=minute)
    job = scheduler.add_job(send_message, trigger)
    print(f"[{datetime.now()}] 更新排程: 每天 {hour}:{minute} 發送訊息")

    last_setting = {"hour": hour, "minute": minute}
    globals()["current_job"] = job
    #print(last_setting)
def update_store():
        load_dotenv()
        Daily_HOST = os.getenv('HRDB_host')
        Daily_password = os.getenv('HRDB_password')
        Daily_uid=os.getenv('HRDB_uid')
        Daily_name=os.getenv('HRDB_name')
        conn = pyodbc.connect(
            f"DRIVER={{ODBC Driver 17 for SQL Server}};"
            f"SERVER={Daily_HOST};"
            f"DATABASE={Daily_name};"
            f"UID={Daily_uid};"
            f"PWD={Daily_password};"
            "Trusted_Connection=no;"
        )
        cursor = conn.cursor()
        
        sql_query = """
        SELECT 
            
            D.[DEP_NAME] AS [部門名稱],
            E.[HECNAME] AS [主管姓名]
        FROM [HRM].[dbo].[HRUSER_DEPT_BAS] AS D
        INNER JOIN [HRM].[dbo].[HRUSER] AS E 
            ON D.[DEP_CHIEF] = E.[EMPID]
        WHERE 
            D.CPNYID = '42756204'
            AND D.DEP_DISABLE != 'Y'
            AND D.DEP_TYPE = '11'
            AND EXISTS (
                SELECT 1 
                FROM [HRM].[dbo].[HRUSER] AS U
                WHERE U.DEPT_NO = D.DEP_NO 
                AND U.STATE = 'A'
            )
        """
        
        cursor.execute(sql_query)
        # 轉換為字典格式： { "部門名稱": "主管姓名" }
        rows = cursor.fetchall()
        sql_store = {row[0].strip(): row[1].strip() for row in rows}
        
        cursor.close()
        conn.close()

        with open("store.json", "r", encoding="utf-8") as f:
            stores = json.load(f)
        updated_stores = []
        processed_names = set()

        # A. 處理「更新」與「刪除」
        # 遍歷舊資料，如果 SQL 裡還有這間店就保留並更新主管，否則就捨棄(刪除)
        for stor in stores:
            name = stor["name"]
            if name in sql_store:
                stor["dept"] = sql_store[name]  # 更新主管姓名
                updated_stores.append(stor)
                processed_names.add(name)
            # else: SQL 沒這間店了，不加入 updated_stores 達成自動刪除

        # B. 處理「新增」
        # 檢查 SQL 裡有哪些店是原本 JSON 裡沒有的
        for dept_name, chief_name in sql_store.items():
            if dept_name not in processed_names:
                new_item = {
                    "value": "",  # 依需求給空字串
                    "name": dept_name,
                    "dept": chief_name
                }
                updated_stores.append(new_item)
        # C. 寫回檔案
        
        with open("store.json", "w", encoding="utf-8") as f:
            json.dump(updated_stores, f, ensure_ascii=False, indent=4)
        
def send_message():
    """發送訊息任務"""
    try:
        update_store()
    except:
        pass
    #day = datetime.today().strftime("%Y-%m-%d")
    day = (datetime.today() - timedelta(days=1)).strftime("%Y-%m-%d")
    with open("permissions.json", "r", encoding="utf-8") as f:
        permissions = json.load(f)
    for per in permissions:
        user_id=per['user_id']
        user_id_LINE=per['LINE']
        user_email=per['email']
        data=getdailydata(user_id,day)
        file_name=excelmake(user_id,day,data,start=5)
 
        if user_email !='':
            Send_EMAIL(user_id,day)
        if user_id_LINE !="":
            send_excel_button(user_id_LINE, file_name,day)#user_id=LINEID
    print(f"[{datetime.now()}] 發送訊息: ")
def excelmake(user_id,day,data,start):#工號 日期資料 完整資料 資料excel期始位置
    with open("permissions.json", "r", encoding="utf-8") as f:
        permission = json.load(f)
    for per in permission:
        if per['user_id']==user_id:
            userdata=per
    user_folder = os.path.join(app.config['FOLDER'], user_id)
    if not os.path.exists(user_folder):
        os.makedirs(user_folder)
    end_date = datetime.strptime(day, "%Y-%m-%d")
    month = end_date.month
    dataday = end_date.day
    date_range_str = f"{month}/1 ~ {month}/{dataday}"
    ytd_range_str=f"1/1 ~ {month}/{dataday}"
    wb = openpyxl.Workbook()
    brand_data = defaultdict(list)
    otherdata=[]
    brand_map = {
        '杏': '杏子豬排',
        '王': '大阪王將',
        '橋': '橋村炸雞',
        '勝': '京都勝牛',
        '雞': '雞三和'
    }
    for row in data:
        store_name = row[0]
        brand = store_name[:1]  # 前兩個字當品牌
        if brand in brand_map:
            rowname = brand_map[brand]
            # if store_name.endswith("Total"):  # 統一處理Total排前面
            #     new_row = row.copy()  # 避免直接改原始 data
            #     new_row[0] = "Total"
            #     brand_data[rowname].insert(0, new_row)
            # else:
            brand_data[rowname].append(row)
        else:
            if store_name.endswith("Total"):
                otherdata.append(row)
    
    for brand ,stores in brand_data.items():
        brand_data[brand].insert(0, otherdata[0])
    
        # if brand=='杏':
        #     rowname='杏子豬排'
        #     brand_data[rowname].append(row)
        # elif brand=='王': 
        #     rowname='大阪王將'
        #     brand_data[rowname].append(row)
        # elif brand=='橋':
        #     rowname='橋村炸雞'
        #     brand_data[rowname].append(row)
        # elif brand=='勝':
        #     rowname='京都勝牛'
        #     brand_data[rowname].append(row)
        # elif brand=='雞': 
        #     rowname='雞三和'   
        #     brand_data[rowname].append(row)
        # else:
        #     otherdata.append(row)
            
        
   
    # 刪掉預設的空白 sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    # 根據品牌放資料
    for brand, data in brand_data.items():
        # print(f"{brand}: {len(data)} 筆")
        # for r in data:
        #     print("   ", r[0])
        # 日期 & 店數
        ws = wb.create_sheet(title=brand)
        ws.merge_cells("A1:B1")
        ws["A1"] = day
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A2"] = "店數"
        ws["B2"] = len(data)-2
        # 標題顏色
        sales_fill = PatternFill("solid", fgColor="800000")   # 暗紅
        tc_fill = PatternFill("solid", fgColor="006666")      # 藍綠
        ta_fill = PatternFill("solid", fgColor="660066")      # 紫色
        header_font = Font(bold=True, color="FFFFFF")
        # 寫 Daily Sales 標題
        ws.merge_cells(f"C{start}:E{start}")
        ws[f"C{start}"] = "Daily Sales"
        ws[f"C{start}"].fill = sales_fill
        ws[f"C{start}"].font = header_font
        ws[f"C{start}"].alignment = Alignment(horizontal="center")
        ws[f"C{start+1}"], ws[f"D{start+1}"], ws[f"E{start+1}"] = "CY", "PY", "Index"
        for col in ["C", "D", "E"]:
            ws[f"{col}{start+1}"].fill = sales_fill
            ws[f"{col}{start+1}"].font = header_font
            ws[f"{col}{start+1}"].alignment = Alignment(horizontal="center")
        # 寫 Daily TC 標題
        ws.merge_cells(f"F{start}:H{start}")
        ws[f"F{start}"] = "Daily TC"
        ws[f"F{start}"].fill = tc_fill
        ws[f"F{start}"].font = header_font
        ws[f"F{start}"].alignment = Alignment(horizontal="center")

        ws[f"F{start+1}"], ws[f"G{start+1}"], ws[f"H{start+1}"] = "CY", "PY", "Index"
        for col in ["F", "G", "H"]:
            ws[f"{col}{start+1}"].fill = tc_fill
            ws[f"{col}{start+1}"].font = header_font
            ws[f"{col}{start+1}"].alignment = Alignment(horizontal="center")

        # 寫 Daily TA 標題
        ws.merge_cells(f"I{start}:K{start}")
        ws[f"I{start}"] = "Daily TA"
        ws[f"I{start}"].fill = ta_fill
        ws[f"I{start}"].font = header_font
        ws[f"I{start}"].alignment = Alignment(horizontal="center")

        ws[f"I{start+1}"], ws[f"J{start+1}"], ws[f"K{start+1}"] = "CY", "PY", "Index"
        for col in ["I", "J", "K"]:
            ws[f"{col}{start+1}"].fill = ta_fill
            ws[f"{col}{start+1}"].font = header_font
            ws[f"{col}{start+1}"].alignment = Alignment(horizontal="center")
        # 寫 MTD Sales 標題
        ws.merge_cells(f"L{start}:N{start}")
        ws[f"L{start}"] = f"MTD Sales({date_range_str})"
        ws[f"L{start}"].fill = sales_fill
        ws[f"L{start}"].font = header_font
        ws[f"L{start}"].alignment = Alignment(horizontal="center")
        ws[f"L{start+1}"], ws[f"M{start+1}"], ws[f"N{start+1}"] = "CY", "PY", "Index"
        for col in ["L", "M", "N"]:
            ws[f"{col}{start+1}"].fill = sales_fill
            ws[f"{col}{start+1}"].font = header_font
            ws[f"{col}{start+1}"].alignment = Alignment(horizontal="center")
        # 寫 MTD TC Sales 標題
        ws.merge_cells(f"O{start}:Q{start}")
        ws[f"O{start}"] = f"MTD TC Sales({date_range_str})"
        ws[f"O{start}"].fill = tc_fill
        ws[f"O{start}"].font = header_font
        ws[f"O{start}"].alignment = Alignment(horizontal="center")
        ws[f"O{start+1}"], ws[f"P{start+1}"], ws[f"Q{start+1}"] = "CY", "PY", "Index"
        for col in ["O", "P", "Q"]:
            ws[f"{col}{start+1}"].fill = tc_fill
            ws[f"{col}{start+1}"].font = header_font
            ws[f"{col}{start+1}"].alignment = Alignment(horizontal="center")
        # 寫 MTD TA Sales 標題
        ws.merge_cells(f"R{start}:T{start}")
        ws[f"R{start}"] = f"MTD TA Sales({date_range_str})"
        ws[f"R{start}"].fill = ta_fill
        ws[f"R{start}"].font = header_font
        ws[f"R{start}"].alignment = Alignment(horizontal="center")
        ws[f"R{start+1}"], ws[f"S{start+1}"], ws[f"T{start+1}"] = "CY", "PY", "Index"
        for col in ["R", "S", "T"]:
            ws[f"{col}{start+1}"].fill = ta_fill
            ws[f"{col}{start+1}"].font = header_font
            ws[f"{col}{start+1}"].alignment = Alignment(horizontal="center")
        # 寫 YTD Sales 標題
        ws.merge_cells(f"U{start}:W{start}")
        ws[f"U{start}"] = f"YTD Sales({ytd_range_str})"
        ws[f"U{start}"].fill = sales_fill
        ws[f"U{start}"].font = header_font
        ws[f"U{start}"].alignment = Alignment(horizontal="center")
        ws[f"U{start+1}"], ws[f"V{start+1}"], ws[f"W{start+1}"] = "CY", "PY", "Index"
        for col in ["U", "V", "W"]:
            ws[f"{col}{start+1}"].fill = sales_fill
            ws[f"{col}{start+1}"].font = header_font
            ws[f"{col}{start+1}"].alignment = Alignment(horizontal="center")
        # 寫 YTD TC Sales 標題
        ws.merge_cells(f"X{start}:Z{start}")
        ws[f"X{start}"] = f"YTD TC Sales({ytd_range_str})"
        ws[f"X{start}"].fill = tc_fill
        ws[f"X{start}"].font = header_font
        ws[f"X{start}"].alignment = Alignment(horizontal="center")
        ws[f"X{start+1}"], ws[f"Y{start+1}"], ws[f"Q{start+1}"] = "CY", "PY", "Index"
        for col in ["X", "Y", "Z"]:
            ws[f"{col}{start+1}"].fill = tc_fill
            ws[f"{col}{start+1}"].font = header_font
            ws[f"{col}{start+1}"].alignment = Alignment(horizontal="center")
        # 寫 YTD TA Sales 標題
        ws.merge_cells(f"AA{start}:AC{start}")
        ws[f"AA{start}"] = f"YTD TA Sales({ytd_range_str})"
        ws[f"AA{start}"].fill = ta_fill
        ws[f"AA{start}"].font = header_font
        ws[f"AA{start}"].alignment = Alignment(horizontal="center")
        ws[f"AA{start+1}"], ws[f"AB{start+1}"], ws[f"AC{start+1}"] = "CY", "PY", "Index"
        for col in ["AA", "AB", "AC"]:
            ws[f"{col}{start+1}"].fill = ta_fill
            ws[f"{col}{start+1}"].font = header_font
            ws[f"{col}{start+1}"].alignment = Alignment(horizontal="center")
        row=start+2
        for r in data:
            ws[f"A{row}"] = r[0]
            ws[f"B{row}"] = r[1]
            ws[f"C{row}"] = r[2]
            ws[f"C{row}"].number_format = "#,##0"
            ws[f"D{row}"] = r[3]
            ws[f"D{row}"].number_format = "#,##0"
            ws[f"E{row}"] = r[4]
            ws[f"E{row}"].number_format = "#,##0.00;-#,##0.00;0"
            ws[f"F{row}"] = r[5]
            ws[f"F{row}"].number_format = "#,##0"
            ws[f"G{row}"] = r[6]
            ws[f"G{row}"].number_format = "#,##0"
            ws[f"H{row}"] = r[7]
            ws[f"H{row}"].number_format = "#,##0.00;-#,##0.00;0"
            ws[f"I{row}"] = r[8]
            ws[f"I{row}"].number_format = "#,##0"
            ws[f"J{row}"] = r[9]
            ws[f"J{row}"].number_format = "#,##0"
            ws[f"K{row}"] = r[10]
            ws[f"K{row}"].number_format = "#,##0.00;-#,##0.00;0"
            ws[f"L{row}"] = r[11]
            ws[f"L{row}"].number_format = "#,##0"
            ws[f"M{row}"] = r[12]
            ws[f"M{row}"].number_format = "#,##0"
            ws[f"N{row}"] = r[13]
            ws[f"N{row}"].number_format = "#,##0.00;-#,##0.00;0"
            ws[f"O{row}"] = r[14]
            ws[f"O{row}"].number_format = "#,##0"
            ws[f"P{row}"] = r[15]
            ws[f"P{row}"].number_format = "#,##0"
            ws[f"Q{row}"] = r[16]
            ws[f"Q{row}"].number_format = "#,##0.00;-#,##0.00;0"
            ws[f"R{row}"] = r[17]
            ws[f"R{row}"].number_format = "#,##0"
            ws[f"S{row}"] = r[18]
            ws[f"S{row}"].number_format = "#,##0"
            ws[f"T{row}"] = r[19]
            ws[f"T{row}"].number_format = "#,##0.00;-#,##0.00;0"
            ws[f"U{row}"] = r[20]
            ws[f"U{row}"].number_format = "#,##0"
            ws[f"V{row}"] = r[21]
            ws[f"V{row}"].number_format = "#,##0"
            ws[f"W{row}"] = r[22]
            ws[f"W{row}"].number_format = "#,##0.00;-#,##0.00;0"
            ws[f"X{row}"] = r[23]
            ws[f"X{row}"].number_format = "#,##0"
            ws[f"Y{row}"] = r[24]
            ws[f"Y{row}"].number_format = "#,##0"
            ws[f"Z{row}"] = r[25]
            ws[f"Z{row}"].number_format = "#,##0.00;-#,##0.00;0"
            ws[f"AA{row}"] = r[26]
            ws[f"AA{row}"].number_format = "#,##0"
            ws[f"AB{row}"] = r[27]
            ws[f"AB{row}"].number_format = "#,##0"
            ws[f"AC{row}"] = r[28]
            ws[f"AC{row}"].number_format = "#,##0.00;-#,##0.00;0"
            row += 1

        # 美化欄寬
        for col in range(1, 30):
            ws.column_dimensions[get_column_letter(col)].width = 15
        ws.column_dimensions[get_column_letter(1)].width = 25    
    wb.save(f"{user_folder}/{day}daily_report.xlsx")
    filename=f"{day}daily_report.xlsx"
    return filename
def Send_EMAIL(user_id,day):#LINEID
    # 郵件內容設定
    sender_email = os.getenv('MAIL')
    password = os.getenv('MAIL_PW')
    with open("permissions.json", "r", encoding="utf-8") as f:
        permission = json.load(f)
    for per in permission:
        if per['user_id']==user_id:
            email=per['email']
    receiver_email=email
    subject = f"{day}日報表"
    filepath = os.path.join(app.config['FOLDER'], user_id, f"{day}daily_report.xlsx")


    # 建立郵件物件
    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = receiver_email
    message["Subject"] = subject
    # 郵件主體
    body_html = f"""
    <html>
    <head><meta charset="utf-8"></head>
    <body>
    <p>附件為{day}日報表再請參考</p>
    </body>
    </html>
    """
    message.attach(MIMEText(body_html, "html"))

    # 加入 Excel 附件
    if os.path.exists(filepath):
        with open(filepath, "rb") as f:
            part = MIMEApplication(f.read(), Name=os.path.basename(filepath))
        part['Content-Disposition'] = f'attachment; filename="{os.path.basename(filepath)}"'
        message.attach(part)
    else:
        print(f"警告：檔案不存在 -> {filepath}")



    try:
        # 建立與 Gmail SMTP 伺服器的連線 (使用 SSL)
        with smtplib.SMTP_SSL("mail.kingza.com.tw", 465) as server:
            if not (isinstance(email, float) and math.isnan(email)):
                server.login(sender_email, password)
                server.sendmail(sender_email, receiver_email, message.as_string())
                print("郵件寄送成功！")

    except Exception as e:
        print(f"發生錯誤：{e}")
def GET_HRdata(user_id):
    load_dotenv()
    HRDB_host = os.getenv('HRDB_host')
    HRDB_password = os.getenv('HRDB_password')
    HRDB_uid=os.getenv('HRDB_uid')
    HRDB_name=os.getenv('HRDB_name')
    conn = pyodbc.connect(
        f"DRIVER={{ODBC Driver 17 for SQL Server}};"
        f"SERVER={HRDB_host};"
        f"DATABASE={HRDB_name};"
        f"UID={HRDB_uid};"
        f"PWD={HRDB_password};"
        "Trusted_Connection=no;"
    )
    cursor = conn.cursor()
    # SUBSTRING(UIDENTID, 2, LEN(UIDENTID) - 1) AS UIDENTID 身分證後九碼
    cursor.execute("SELECT EMPID, HECNAME ,EMAIL FROM HRM.dbo.HRUSER WHERE EMPID = ?", (user_id,))
    row = cursor.fetchone()
    

    if row and any(row):
        return {'ID': row[0], 'name': row[1], 'email': row[2]}
    else:
        return None
@app.route("/data")
def index():
    with open("permissions.json", "r", encoding="utf-8") as f:
        permissions = json.load(f)
    with open("store.json", "r", encoding="utf-8") as f:
        stores = json.load(f)
    alldata={}
    stores.append({'value': 'all', 'name': '全門市', 'dept': '全門市'})
   

    return render_template("index.html", permissions=permissions,stores=stores,settings=settings)
@app.route("/adduser", methods=['POST'])
def adduser():
    with open("permissions.json", "r", encoding="utf-8") as f:
        permissions = json.load(f)
    data = request.get_json()
    permissions.append({
        "departments": data.get("storeValues", []),
        "email": data.get("email"),
        "user_id": data.get("user"),
        "name": data.get("name"),
        "LINE": data.get("LINE")
    })
    with open("permissions.json", "w", encoding="utf-8") as f:
        json.dump(permissions, f, ensure_ascii=False, indent=4)
    return jsonify({"success": True, "message": "資料已儲存"})
@app.route("/edituser", methods=['POST'])
def edituser():
    with open("permissions.json", "r", encoding="utf-8") as f:
        permissions = json.load(f)
    data = request.get_json()
    for per in permissions:
        if per['user_id']==data['editUser']:
            per['name']=data['editName']
            per['departments']=data['editStore']
            per['email']=data['editEmail']
            per['LINE']=data['editLINE']
            break
    with open("permissions.json", "w", encoding="utf-8") as f:
        json.dump(permissions, f, ensure_ascii=False, indent=4)
    return jsonify({"success": True, "message": "資料已儲存"})
@app.route("/deletuser", methods=['POST'])
def deletuser():
    with open("permissions.json", "r", encoding="utf-8") as f:
        permissions = json.load(f)
    data = request.get_json()
    newdata=[]
    for per in permissions:
        if per['user_id']==data['userid']:
            pass
        else:
            newdata.append(per)
    with open("permissions.json", "w", encoding="utf-8") as f:
        json.dump(newdata, f, ensure_ascii=False, indent=4)
    return jsonify({"success": True, "message": "資料已儲存"})
@app.route("/home")
def home():
    return render_template("home.html" )
@app.route("/store")
def store():
    with open("store.json", "r", encoding="utf-8") as f:
        stores = json.load(f)
    return render_template("store.html",stores=stores,settings=settings)
@app.route("/addstore", methods=['POST'])
def addstore():
    with open("store.json", "r", encoding="utf-8") as f:
        store = json.load(f)
    data = request.get_json()
    store.append({
        "value": data.get("value"),
        "name": data.get("name"),
        "dept": data.get("dept")
    })
    with open("store.json", "w", encoding="utf-8") as f:
        json.dump(store, f, ensure_ascii=False, indent=4)
    return jsonify({"success": True, "message": "資料已儲存"})
@app.route("/deletstore", methods=['POST'])
def deletstore():
    with open("store.json", "r", encoding="utf-8") as f:
        store = json.load(f)
    data = request.get_json()
    newdata=[]
    for per in store:
        if per['value']==data['value']:
            pass
        else:
            newdata.append(per)
    with open("store.json", "w", encoding="utf-8") as f:
        json.dump(newdata, f, ensure_ascii=False, indent=4)
    return jsonify({"success": True, "message": "資料已儲存"})
@app.route("/editstore", methods=['POST'])
def editstore():
    with open("store.json", "r", encoding="utf-8") as f:
        store = json.load(f)
    data = request.get_json()
    for per in store:
        if per['name']==data['name']:
            per['value']=data['value']
            per['dept']=data['dept']
            break
    with open("store.json", "w", encoding="utf-8") as f:
        json.dump(store, f, ensure_ascii=False, indent=4)
    return jsonify({"success": True, "message": "資料已儲存"})
@app.route("/setting")
def setting():
    with open("settings.json", "r", encoding="utf-8") as f:
        setting = json.load(f)
    return render_template("setting.html",setting=setting,settings=settings)
@app.route("/editsetting", methods=['POST'])
def editsetting():
    with open("settings.json", "r", encoding="utf-8") as f:
        setting = json.load(f)
    data = request.get_json()
    for per in setting:
        per['hour']=data['hour']
        per['minute']=data['minute']
        per['ngrokid']=data['ngrokid']
        break
    with open("settings.json", "w", encoding="utf-8") as f:
        json.dump(setting, f, ensure_ascii=False, indent=4)
    return jsonify({"success": True, "message": "資料已儲存"})
@app.route("/files/<user_id>/<path:filename>")
def serve_file(user_id,filename):
    folder_path = os.path.join(app.config['FOLDER'], user_id)
    print(filename)
    return send_from_directory(folder_path, filename, as_attachment=True)
@app.route("/png/<path:filename>")
def png_file(filename):
    return send_from_directory(app.config['PNG'], filename, as_attachment=True)
#================LINE WEBHOOK=====================
@app.route("/callback", methods=['POST'])
def callback():
    signature = request.headers['X-Line-Signature']
    body = request.get_data(as_text=True)

    try:
        handler.handle(body, signature)
        
    except InvalidSignatureError:
        abort(400)

    return 'OK'
# 發送檔案下載連結
def send_excel_button(user_id, file_name,day):
    with open("settings.json", "r", encoding="utf-8") as f:
        setting = json.load(f)
    setting = setting[0]
    
    with open("permissions.json", "r", encoding="utf-8") as f:
        permission = json.load(f)
    for per in permission:
        if per['LINE']==user_id:
            ID=per['user_id']
    file_url = f"https://{setting['ngrokid']}/files/{ID}/{file_name}"
    buttons_template = ButtonsTemplate(
        thumbnail_image_url=f"https://{setting['ngrokid']}/png/logo.png",
        title="日報表",
        text=day,
        actions=[
            URIAction(label="Download", uri=file_url),
        ]
    )
    message = TemplateMessage(
        alt_text="ButtonsTemplate",
        template=buttons_template
    )
    line_bot_api.push_message(
        PushMessageRequest(
            to=user_id,
            messages=[message]
        )
    )
# ====== 使用者傳訊息事件 ======
@handler.add(MessageEvent, message=TextMessageContent)
def handle_message(event):
    user_id = event.source.user_id
    user_text = event.message.text.strip()

    print(f"收到來自 {user_id} 的訊息: {user_text}")
    if user_text[:2] == "工號":
        rest = user_text[2:]          # 取 "工號" 後面的字
        
        if rest[0].lower() == "a" or rest[0].lower() == "A":
            if rest[0].lower() == "a":
                rest = "A" + rest[1:]
            result = rest
            HR=GET_HRdata(result)
            if HR !=None:
                line_index = None
                user_index = None
                email=HR['email']
                name=HR['name']
                with open("permissions.json", "r", encoding="utf-8") as f:
                    permissions = json.load(f)
                for i, per in enumerate(permissions):
                    if per.get('LINE') == user_id:
                        line_index = i
                    if per.get('user_id') == result:
                        user_index = i
                if line_index is not None and user_index is not None:
                    if line_index == user_index:
                        text=f'已有帳號:\n工號:{result}\n名稱:{name}\n電子郵件:{email}\n如有任何問題請洽管理員'
                    elif line_index != user_index:
                        # 兩筆分開 → 合併成一筆
                        per_line = permissions[line_index]
                        per_user = permissions[user_index]

                        # 合併資訊 (以 LINE 為主，補上 user)
                        per_line['user_id'] = result  

                        # 刪掉 LINE 那筆
                        del permissions[user_index]
                        text=f'已有帳號:\n工號:{result}\n名稱:{name}\n電子郵件:{email}\n如有任何問題更改請洽管理員'

                elif line_index is not None:
                    # 只有 LINE 存在 → 補 user_id
                    permissions[line_index]['user_id'] = result
                    permissions[line_index]['email'] = email
                    permissions[line_index]['name'] = name
                    permissions[line_index]['departments'] = []
                    text=f'已更新帳號:\n工號:{result}\n名稱:{name}\n電子郵件:{email}\n如有任何問題請洽管理員'

                elif user_index is not None:
                    # 只有 user_id 存在 → 補 LINE
                    permissions[user_index]['LINE'] = user_id
                    hname=permissions[user_index]['name'] 
                    hemail=permissions[user_index]['email'] 
                        
                    text=f'已更新帳號:\n工號:{result}\n名稱:{hname}\n電子郵件:{hemail}\n如有任何問題請洽管理員'
                    print(result)
                else:
                    # 兩個都沒有 → 新增一筆
                    permissions.append({
                        "departments" : [],
                        "email" :email,
                        "user_id": result,
                        "name":name,
                        "LINE": user_id
                    })
                    text=f'已新增帳號:\n工號:{result}\n名稱:{name}\n電子郵件:{email}\n如有任何問題更改請洽管理員'

                with open("permissions.json", "w", encoding="utf-8") as f:
                    json.dump(permissions, f, ensure_ascii=False, indent=4)
            else:
                 text=f'查無此工號!!\n\n請提供管理員以下訊息供查詢使用\n\n工號:{result}\nID:{user_id}'

            # 回覆訊息
            line_bot_api.reply_message_with_http_info(
                ReplyMessageRequest(
                    reply_token=event.reply_token,
                    messages=[TextMessage(text=text)]
                    )
            )
    elif user_text[:2] == "資料":
        date_str=user_text[2:10]    
        with open("permissions.json", "r", encoding="utf-8") as f:
            permissions = json.load(f)
        for per in permissions:
            if user_id == per['LINE']:
                user_id=per['user_id']
                user_id_LINE=per['LINE']
                day = datetime.strptime(date_str, "%Y%m%d").strftime("%Y-%m-%d")
                data=getdailydata(user_id,day)
                file_name=excelmake(user_id,day,data,start=5)
                send_excel_button(user_id_LINE, file_name,day)
                break

      

# ====== 使用者加好友事件 (FollowEvent) ======
# @handler.add(FollowEvent)
# def handle_follow(event):
#     user_id = event.source.user_id
#     print("新加入的使用者 ID:", user_id)

#     # 可以回覆一則歡迎訊息
#     line_bot_api.reply_message(
#         event.reply_token,
#         TextSendMessage(text=f"歡迎加入！你的ID是 {user_id}")
#     )







# # ===== LINE 推送功能 =====
# def push_excel_link():
#     for uid, roles in permissions.items():
#         for role in roles:
#             file_name = f"{role}_report.xlsx"
#             file_url = f"http://your-public-ip:5000/files/{file_name}"  # 需要改成你 public IP 或 ngrok
#             try:
#                 line_bot_api.push_message(uid, TextSendMessage(text=f"您的檔案下載連結：{file_url}"))
#             except Exception as e:
#                 print(f"推送失敗 {uid}: {e}")

# # ===== APScheduler 排程 =====
# scheduler = BackgroundScheduler()
# scheduler.add_job(
#     push_excel_link,
#     'cron',
#     hour=settings["hour"],
#     minute=settings["minute"]
# )
# scheduler.start()

# ===== Flask 啟動 =====
scheduler = BackgroundScheduler()
current_job = None
scheduler.add_job(update_job, 'interval', minutes=1)
scheduler.start()
# day = datetime.today().strftime("%Y-%m-%d")
# data=getdailydata('A14176',day)
# excelmake('A14176',day,data,5)
#day = datetime.today().strftime("%Y-%m-%d")

day = (datetime.today() - timedelta(days=1)).strftime("%Y-%m-%d")
day='2026-03-09'
data=getdailydata("A14176",day)
excelmake('A14176',day,data,5)
update_store()
#使用FLASK啟動須解除，目前以Gunicorn啟動
if __name__ == "__main__":

    app.run(host="0.0.0.0", port=8018)
